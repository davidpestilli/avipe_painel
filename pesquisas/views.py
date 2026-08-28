from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

from django.conf import settings
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render

from .analytics import buscar_observabilidade
from .services import (
    buscar_metricas,
    buscar_registro_detalhe,
    consultar_registros,
    exportar_processos,
    exportar_registros,
    listar_ambientes_disponiveis,
    listar_siglas_orgaos,
    listar_ultimos_registros,
    listar_usuarios_logados,
    obter_info_banco,
    resolver_ambiente_request,
)

_UTC = ZoneInfo("UTC")
_SAO_PAULO = ZoneInfo("America/Sao_Paulo")
_EXPORT_LIMIT = 1000


def _contexto_base() -> dict[str, Any]:
    app_js = Path(settings.BASE_DIR) / "frontend" / "dist" / "assets" / "app.js"
    app_css = Path(settings.BASE_DIR) / "frontend" / "dist" / "assets" / "app.css"
    asset_version = "dev"
    if app_js.exists():
        asset_version = str(int(app_js.stat().st_mtime))
    elif app_css.exists():
        asset_version = str(int(app_css.stat().st_mtime))

    return {
        "titulo_app": "Watcher AVIPE",
        "asset_version": asset_version,
    }


def react_app(request: HttpRequest) -> HttpResponse:
    return render(request, "pesquisas/react_app.html", _contexto_base())


def health(request: HttpRequest) -> JsonResponse:
    return JsonResponse({"status": "ok"})


def _ajustar_datas_para_fuso_local(valor: Any) -> Any:
    if isinstance(valor, datetime):
        base = valor.replace(tzinfo=_UTC) if valor.tzinfo is None else valor
        return base.astimezone(_SAO_PAULO)
    return valor


def _normalizar_registro_datas(registro: dict[str, Any]) -> dict[str, Any]:
    return {
        chave: _ajustar_datas_para_fuso_local(valor)
        for chave, valor in registro.items()
    }


def _serializar_registro(registro: dict[str, Any]) -> dict[str, Any]:
    normalizado = _normalizar_registro_datas(registro)
    serializado: dict[str, Any] = {}
    for chave, valor in normalizado.items():
        if isinstance(valor, datetime):
            serializado[chave] = valor.isoformat()
        else:
            serializado[chave] = valor
    return serializado


def api_dashboard(request: HttpRequest) -> JsonResponse:
    ambiente = resolver_ambiente_request(request)
    try:
        return JsonResponse(
            {
                "metricas": buscar_metricas(ambiente),
                "info_banco": obter_info_banco(ambiente),
                "ultimos_registros": [
                    _serializar_registro(item)
                    for item in listar_ultimos_registros(ambiente=ambiente)
                ],
                "siglas_orgaos": listar_siglas_orgaos(ambiente),
            }
        )
    except Exception as exc:  # pragma: no cover
        return JsonResponse({"erro": str(exc)}, status=500)


def api_observabilidade(request: HttpRequest) -> JsonResponse:
    periodo = request.GET.get("periodo", "today")
    ambiente = resolver_ambiente_request(request)
    try:
        return JsonResponse(buscar_observabilidade(periodo, ambiente))
    except Exception as exc:  # pragma: no cover
        return JsonResponse({"erro": str(exc)}, status=500)


def api_configuracoes(request: HttpRequest) -> JsonResponse:
    ambiente = resolver_ambiente_request(request)
    try:
        return JsonResponse(
            {
                "ambiente_ativo": ambiente,
                "ambientes": listar_ambientes_disponiveis(),
                "info_banco": obter_info_banco(ambiente),
            }
        )
    except Exception as exc:  # pragma: no cover
        return JsonResponse({"erro": str(exc)}, status=500)


def api_lista_pesquisas(request: HttpRequest) -> JsonResponse:
    ambiente = resolver_ambiente_request(request)
    filtros = _extrair_filtros_pesquisa(request)
    pagina = max(1, int(request.GET.get("pagina", "1") or "1"))
    try:
        paginacao = _normalizar_paginacao(
            consultar_registros(filtros=filtros, pagina=pagina, ambiente=ambiente)
        )
        query_base = {chave: valor for chave, valor in filtros.items() if valor}
        return JsonResponse(
            {
                "ambiente_ativo": ambiente,
                "filtros": filtros,
                "siglas_orgaos": listar_siglas_orgaos(ambiente),
                "usuarios_logados": listar_usuarios_logados(ambiente),
                "paginacao": {
                    "itens": [_serializar_registro(item) for item in paginacao.itens],
                    "pagina": paginacao.pagina,
                    "por_pagina": paginacao.por_pagina,
                    "total": paginacao.total,
                    "total_processos": paginacao.total_processos,
                    "total_paginas": paginacao.total_paginas,
                    "tem_anterior": paginacao.tem_anterior,
                    "tem_proxima": paginacao.tem_proxima,
                    "anterior_url": f"/api/pesquisas/?{urlencode({**query_base, 'pagina': paginacao.pagina - 1})}" if paginacao.tem_anterior else "",
                    "proxima_url": f"/api/pesquisas/?{urlencode({**query_base, 'pagina': paginacao.pagina + 1})}" if paginacao.tem_proxima else "",
                },
            }
        )
    except Exception as exc:  # pragma: no cover
        return JsonResponse({"erro": str(exc)}, status=500)


def api_detalhe_pesquisa(request: HttpRequest) -> JsonResponse:
    ambiente = resolver_ambiente_request(request)
    registro_id = request.GET.get("id", "")
    if not registro_id:
        return JsonResponse({"erro": "Informe o id do registro para abrir o detalhe."}, status=400)

    try:
        registro = buscar_registro_detalhe(int(registro_id), ambiente)
    except Exception as exc:  # pragma: no cover
        return JsonResponse({"erro": str(exc)}, status=500)

    if not registro:
        return JsonResponse({"erro": "Registro nao encontrado com os parametros informados."}, status=404)

    return JsonResponse({"ambiente_ativo": ambiente, "registro": _serializar_registro(registro)})


def api_exportar_pesquisas(request: HttpRequest) -> HttpResponse:
    ambiente = resolver_ambiente_request(request)
    filtros = _extrair_filtros_pesquisa(request)
    modo = request.GET.get("modo", "agrupada").strip().lower()

    try:
        workbook_factory = _obter_workbook_factory()
        if modo == "linhas":
            registros = [
                _normalizar_registro_datas(item)
                for item in exportar_registros(filtros=filtros, limite=_EXPORT_LIMIT, ambiente=ambiente)
            ]
            response = _montar_resposta_excel_registros(registros, workbook_factory)
            nome_arquivo = "pesquisas_registros.xlsx"
        else:
            processos = [
                _normalizar_registro_datas(item)
                for item in exportar_processos(filtros=filtros, limite=_EXPORT_LIMIT, ambiente=ambiente)
            ]
            response = _montar_resposta_excel_processos(processos, workbook_factory)
            nome_arquivo = "pesquisas_processos.xlsx"
    except ModuleNotFoundError as exc:  # pragma: no cover
        if exc.name == "openpyxl":
            return JsonResponse(
                {"erro": "A exportacao para Excel requer a biblioteca openpyxl instalada no ambiente virtual."},
                status=500,
            )
        return JsonResponse({"erro": str(exc)}, status=500)
    except Exception as exc:  # pragma: no cover
        return JsonResponse({"erro": str(exc)}, status=500)

    response["Content-Disposition"] = f'attachment; filename="{nome_arquivo}"'
    return response


def _extrair_filtros_pesquisa(request: HttpRequest) -> dict[str, str]:
    return {
        "nuprocesso": request.GET.get("nuprocesso", ""),
        "cpf": request.GET.get("cpf", ""),
        "sig_orgao": request.GET.get("sig_orgao", ""),
        "usuario_logado": request.GET.get("usuario_logado", ""),
        "data_insercao_status": request.GET.get("data_insercao_status", ""),
        "data_insercao_inicio": request.GET.get("data_insercao_inicio", ""),
        "data_insercao_fim": request.GET.get("data_insercao_fim", ""),
        "data_processamento_status": request.GET.get("data_processamento_status", ""),
        "data_processamento_inicio": request.GET.get("data_processamento_inicio", ""),
        "data_processamento_fim": request.GET.get("data_processamento_fim", ""),
        "processado": request.GET.get("processado", ""),
        "juntado": request.GET.get("juntado", ""),
    }


def _obter_workbook_factory():
    from openpyxl import Workbook

    return Workbook


def _montar_resposta_excel_registros(registros: list[dict[str, Any]], workbook_factory) -> HttpResponse:
    workbook = workbook_factory()
    sheet = workbook.active
    sheet.title = "Registros"
    sheet.append(
        [
            "ID",
            "Processo",
            "CPF/CNPJ",
            "Orgao",
            "Usuario",
            "IP cliente",
            "Inserido em",
            "Incluido no localizador em",
            "Processado em",
            "Processado",
            "Juntado",
            "Localizado",
        ]
    )

    for item in registros:
        sheet.append(
            [
                item.get("id"),
                _formatar_celula_texto(item.get("nuprocesso")),
                _formatar_celula_texto(item.get("cpf")),
                item.get("sig_orgao"),
                item.get("usuario_logado"),
                item.get("ip_cliente"),
                _formatar_celula_data(item.get("data_insercao")),
                _formatar_celula_data(item.get("data_inclusao_localizador")),
                _formatar_celula_data(item.get("data_processamento")),
                _formatar_celula_booleano(item.get("processado")),
                _formatar_celula_booleano(item.get("juntado")),
                _formatar_celula_booleano(item.get("localizado")),
            ]
        )

    _formatar_planilha(sheet, text_columns=("B", "C"))
    return _serializar_workbook(workbook)


def _montar_resposta_excel_processos(processos: list[dict[str, Any]], workbook_factory) -> HttpResponse:
    workbook = workbook_factory()
    sheet = workbook.active
    sheet.title = "Processos"
    sheet.append(
        [
            "Processo",
            "Quantidade de registros",
            "Orgao",
            "Usuario",
            "Inserido em",
            "Incluido no localizador em",
            "Processado em",
            "Processado",
            "Juntado",
        ]
    )

    for item in processos:
        sheet.append(
            [
                _formatar_celula_texto(item.get("nuprocesso")),
                item.get("total_registros"),
                item.get("sig_orgao"),
                item.get("usuario_logado"),
                _formatar_celula_data(item.get("data_insercao")),
                _formatar_celula_data(item.get("data_inclusao_localizador")),
                _formatar_celula_data(item.get("data_processamento")),
                _formatar_celula_booleano(item.get("processado")),
                _formatar_celula_booleano(item.get("juntado")),
            ]
        )

    _formatar_planilha(sheet, text_columns=("A",))
    return _serializar_workbook(workbook)


def _serializar_workbook(workbook) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _formatar_celula_data(valor: Any) -> str:
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M:%S")
    return "" if valor in (None, "") else str(valor)


def _formatar_celula_booleano(valor: Any) -> str:
    if valor in (True, 1, "1"):
        return "Sim"
    if valor in (False, 0, "0"):
        return "Nao"
    return ""


def _formatar_celula_texto(valor: Any) -> str:
    return "" if valor in (None, "") else str(valor)


def _formatar_planilha(sheet, text_columns: tuple[str, ...] = ()) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(fill_type="solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    centered = Alignment(horizontal="center", vertical="center")
    left_aligned = Alignment(horizontal="left", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered

    for column_name in text_columns:
        for cell in sheet[column_name]:
            cell.number_format = "@"

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = left_aligned if cell.column == 1 else centered

    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            if len(text) > max_length:
                max_length = len(text)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 32)


def _normalizar_paginacao(paginacao):
    paginacao.itens = [_normalizar_registro_datas(item) for item in paginacao.itens]
    return paginacao
